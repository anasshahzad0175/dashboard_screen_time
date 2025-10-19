import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image as PILImage, ImageDraw, ImageFont
import os


pd.set_option('future.no_silent_downcasting', True)

plt.rcParams['axes.facecolor'] = 'white'
plt.rcParams['figure.facecolor'] = 'white'


os.makedirs("tmp_charts", exist_ok=True)


df = pd.read_csv("data.csv")

gender_map = {1: "Female", 2: "Male"}
occupation_map = {1: "Employed", 2: "Student", 3: "Self-employed", 4: "Unemployed", 5: "Retired"}
work_mode_map = {1: "Remote", 2: "Hybrid", 3: "In-person"}
sleep_quality_map = {1: "Very Poor", 2: "Poor", 3: "Good", 4: "Excellent"}

df["gender"] = df["gender"].map(gender_map)
df["occupation"] = df["occupation"].map(occupation_map)
df["work_mode"] = df["work_mode"].map(work_mode_map)
df["sleep_quality_1_5"] = df["sleep_quality_1_5"].map(sleep_quality_map)


for col in ["screen_time_hours", "sleep_hours", "stress_level_0_10", "work_screen_hours", "leisure_screen_hours"]:
    df[col] = df[col].round(2)


avg_work_screen = df["work_screen_hours"].mean().round(2)
avg_leisure_screen = df["leisure_screen_hours"].mean().round(2)
avg_sleep_quality_mode = df["sleep_quality_1_5"].mode()[0]

avg_stress = df["stress_level_0_10"].mean().round(2)
avg_total_screen = df["screen_time_hours"].mean().round(2)


def create_kpi_card(title, value, subtitle, filename, start_color, end_color):
    """Create a gradient KPI card as an image."""
    w, h = 400, 150
    img = PILImage.new("RGB", (w, h), color=start_color)
    draw = ImageDraw.Draw(img)

    for y in range(h):
        ratio = y / h
        r = int(start_color[0] * (1 - ratio) + end_color[0] * ratio)
        g = int(start_color[1] * (1 - ratio) + end_color[1] * ratio)
        b = int(start_color[2] * (1 - ratio) + end_color[2] * ratio)
        draw.line([(0, y), (w, y)], fill=(r, g, b))

    try:
        title_font = ImageFont.truetype("arial.ttf", 18)
        value_font = ImageFont.truetype("arialbd.ttf", 32)
        subtitle_font = ImageFont.truetype("arial.ttf", 14)
    except:
        title_font = ImageFont.load_default()
        value_font = ImageFont.load_default()
        subtitle_font = ImageFont.load_default()

    value_text = str(value)
    bbox_val = draw.textbbox((0, 0), value_text, font=value_font)
    w_val = bbox_val[2] - bbox_val[0]
    h_val = bbox_val[3] - bbox_val[1]

    bbox_title = draw.textbbox((0, 0), title, font=title_font)
    w_title = bbox_title[2] - bbox_title[0]

    draw.text(((w - w_title) / 2, 20), title, fill="white", font=title_font)
    draw.text(((w - w_val) / 2, 50), value_text, fill="white", font=value_font)
    draw.text((20, 110), subtitle, fill="white", font=subtitle_font)

    img.save(filename)
    return filename


kpi_files = []
kpi_files.append(create_kpi_card("Avg Work Screen Time", f"{avg_work_screen:.2f} hrs",
                                 "Work-related screen usage", "tmp_charts/kpi1.png",
                                 (84,142,212), (154,196,255)))
kpi_files.append(create_kpi_card("Avg Leisure Screen Time", f"{avg_leisure_screen:.2f} hrs",
                                 "Entertainment & social", "tmp_charts/kpi2.png",
                                 (77,175,124), (161,225,188)))
kpi_files.append(create_kpi_card("Most Common Sleep Quality", avg_sleep_quality_mode,
                                 "Most frequent sleep rating", "tmp_charts/kpi3.png",
                                 (255,163,72), (255,207,140)))
kpi_files.append(create_kpi_card("Avg Stress Level", f"{avg_stress:.2f}/10",
                                 "Reported stress level", "tmp_charts/kpi4.png",
                                 (230,90,90), (255,153,153)))
kpi_files.append(create_kpi_card("Avg Total Screen Time", f"{avg_total_screen:.2f} hrs",
                                 "Daily screen exposure", "tmp_charts/kpi5.png",
                                 (100,100,255), (160,180,255)))


plt.figure(figsize=(6,4))
df.groupby("occupation")["sleep_hours"].mean().plot(kind="bar", color="#558ED5")
plt.title("Average Sleeping Hours by Occupation")
plt.xlabel("Occupation"); plt.ylabel("Hours")
plt.tight_layout(); plt.savefig("tmp_charts/chart1.png"); plt.close()

plt.figure(figsize=(6,4))
means = df.groupby("work_mode")["sleep_hours"].mean().round(2)
bars = plt.bar(means.index, means.values, color="#70AD47")

for bar, value in zip(bars, means.values):
    plt.text(
        bar.get_x() + bar.get_width()/2,
        bar.get_height() + 0.05,
        f"{value:.2f}",
        ha="center", va="bottom", fontsize=10, fontweight="bold", color="black"
    )

plt.title("Average Sleeping Hours by Work Mode")
plt.xlabel("Work Mode")
plt.ylabel("Hours")
plt.ylim(0, means.max() + 1) 
plt.tight_layout()
plt.savefig("tmp_charts/chart2.png")
plt.close()


plt.figure(figsize=(6,4))
df.boxplot(column="screen_time_hours", by="sleep_quality_1_5", patch_artist=True,
           boxprops=dict(facecolor="#409AB5"), 
           medianprops=dict(color="#E1F3F8", linewidth=1),
           whiskerprops=dict(color="gray"),
           capprops=dict(color="gray"),
           flierprops=dict(marker='o', color='darkblue', alpha=0.5))
plt.title("Screen Time Distribution by Sleep Quality")
plt.xlabel("Sleep Quality", fontsize=11)
plt.ylabel("Screen Time (hours)", fontsize=11)
plt.suptitle("")
plt.tight_layout(); plt.savefig("tmp_charts/chart3.png"); plt.close()

plt.figure(figsize=(6,4))
plt.scatter(df["screen_time_hours"], df["stress_level_0_10"], c="skyblue", alpha=0.6, edgecolors="gray")
plt.title("Screen Time vs Stress Level")
plt.xlabel("Screen Time (hrs)"); plt.ylabel("Stress Level (0-10)")
plt.tight_layout(); plt.savefig("tmp_charts/chart4.png"); plt.close()


wb = Workbook()
ws = wb.active
ws.title = "Lifestyle Dashboard"


ws.merge_cells("A1:J2")
cell = ws["A1"]
cell.value = "Lifestyle & Screen Time Premium Dashboard"
cell.font = Font(size=22, bold=True, color="1F4E79")
cell.alignment = Alignment(horizontal="center", vertical="center")


x_pos = 1
for file in kpi_files:
    img = Image(file)
    ws.add_image(img, f"A{3 + (x_pos-1)*10}")
    x_pos += 1

chart_files = ["tmp_charts/chart1.png", "tmp_charts/chart2.png", "tmp_charts/chart3.png", "tmp_charts/chart4.png"]
positions = ["A60", "F60", "A95", "F95"]
for pos, file in zip(positions, chart_files):
    ws.add_image(Image(file), pos)

ws["A130"] = "📊 Key Insights:"
ws["A130"].font = Font(bold=True, size=16, color="1F4E79")
ws["A131"] = (
    "- Most people sleep ~7 hours/day.\n"
    "- Higher screen time leads to lower sleep quality.\n"
    "- Stress increases with screen exposure.\n"
    "- Work mode & job type have limited effect on sleep.\n"
    "- Healthy digital habits improve lifestyle balance."
)
ws["A131"].alignment = Alignment(wrap_text=True)


wb.save("new_Dashboard.xlsx")
print("✅ Premium Excel dashboard saved as 'Lifestyle_Premium_Light_Dashboard.xlsx'")
